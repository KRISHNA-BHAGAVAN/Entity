
import io
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def generate_report_excel(
    columns: List[str], 
    data: List[Dict[str, Any]], 
    time_range_label: str
) -> bytes:
    """
    Generates an Excel report for the given data.
    
    Args:
        columns: List of column headers (ordered).
        data: List of dictionaries, each representing a row (event). Keys must match columns.
        time_range_label: Label for the report (e.g., "Weekly Report").
        
    Returns:
        Bytes of the generated .xlsx file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Events Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    
    # 1. Title Row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    title_cell = ws.cell(row=1, column=1, value=f"Consolidated Report: {time_range_label}")
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = center_align
    
    # 2. Header Row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # 3. Data Rows
    for row_idx, row_data in enumerate(data, 3):
        for col_idx, col_name in enumerate(columns, 1):
            val = row_data.get(col_name)
            
            # Handle list values (e.g. from multiple references) by joining them
            if isinstance(val, list):
                val = ", ".join(str(v) for v in val)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            
    # 4. Auto-size columns (approximated)
    for col_idx, col_name in enumerate(columns, 1):
        # Base width on header length
        max_length = len(col_name)
        
        # Check first few rows of data to adjust width
        for i in range(min(len(data), 10)):
            val = data[i].get(col_name)
            if val:
                max_length = max(max_length, len(str(val)))
        
        adjusted_width = min(max_length + 2, 50) # Cap width at 50 chars
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = adjusted_width

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
